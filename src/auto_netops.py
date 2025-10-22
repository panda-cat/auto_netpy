#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import netmiko
import openpyxl
import argparse
import os
import platform
import datetime
import sys
import re
import uuid
import tempfile
import time
from typing import List, Dict, Optional, Tuple, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import encodings.idna
from tqdm import tqdm
from netmiko import NetmikoTimeoutException, NetmikoAuthenticationException
from netmiko.ssh_dispatcher import CLASS_MAPPER

# ==============================================================================
# 核心配置与环境
# ==============================================================================
os.environ["NO_COLOR"] = "1"
write_lock = Lock()
# 优化: 限制线程数到更安全的范围 (IO密集型任务: CPU * 5 倍，上限 128)
DEFAULT_MAX_THREADS = 128
DEFAULT_THREADS = min(DEFAULT_MAX_THREADS, max(4, (os.cpu_count() or 4) * 5))

# 动态获取所有支持的设备类型
SUPPORTED_DEVICE_TYPES = set(CLASS_MAPPER.keys())

# 设备类型别名映射（常用别名到标准类型）
DEVICE_TYPE_ALIASES = {
    # Cisco设备
    'cisco': 'cisco_ios', 'cisco_switch': 'cisco_ios', 'cisco_router': 'cisco_ios', 'cisco_catalyst': 'cisco_ios',
    'nexus': 'cisco_nxos', 'cisco_nexus': 'cisco_nxos', 'asa': 'cisco_asa', 'cisco_firewall': 'cisco_asa',
    'ios_xe': 'cisco_xe', 'ios_xr': 'cisco_xr', 'cisco_wlc': 'cisco_wlc_ssh',
    
    # Huawei设备
    'huawei': 'huawei', 'huawei_switch': 'huawei', 'huawei_router': 'huawei', 'huawei_firewall': 'huawei',
    'vrp': 'huawei', 'vrpv8': 'huawei_vrpv8', 'huawei_vrp': 'huawei_vrpv8',
    
    # HP/H3C/Aruba设备
    'hp': 'hp_comware', 'hp_switch': 'hp_comware', 'comware': 'hp_comware', 'h3c': 'hp_comware',
    'procurve': 'hp_procurve', 'hp_procurve_switch': 'hp_procurve', 'aruba': 'aruba_os',
    
    # Juniper设备
    'juniper': 'juniper', 'junos': 'juniper', 'juniper_switch': 'juniper', 'juniper_router': 'juniper',
    'juniper_firewall': 'juniper_screenos', 'srx': 'juniper_screenos',
    
    # Fortinet/PaloAlto设备
    'fortinet': 'fortinet', 'fortigate': 'fortinet', 'fortios': 'fortinet', 'fortinet_firewall': 'fortinet',
    'paloalto': 'paloalto_panos', 'panos': 'paloalto_panos', 'pa': 'paloalto_panos', 'paloalto_firewall': 'paloalto_panos',
    
    # Dell设备
    'dell': 'dell_force10', 'force10': 'dell_force10', 'dell_powerconnect': 'dell_powerconnect',
    'dell_os6': 'dell_os6', 'dell_os9': 'dell_os9', 'dell_os10': 'dell_os10',
    
    # Extreme设备
    'extreme': 'extreme', 'extreme_switch': 'extreme', 'extreme_exos': 'extreme_exos',
    'exos': 'extreme_exos', 'extreme_wing': 'extreme_wing',
    
    # Ruckus/Brocade设备
    'ruckus': 'ruckus_fastiron', 'ruckus_switch': 'ruckus_fastiron', 'ruckus_icx': 'ruckus_fastiron',
    'fastiron': 'ruckus_fastiron', 'brocade': 'ruckus_fastiron',
    
    # Mikrotik设备
    'mikrotik': 'mikrotik_routeros', 'routeros': 'mikrotik_routeros', 'mikrotik_router': 'mikrotik_routeros',
    
    # 其他设备
    'alcatel': 'alcatel_aos', 'nokia': 'nokia_sros', 'sros': 'nokia_sros', 'avaya': 'avaya_ers',
    'allied_telesis': 'allied_telesis_awplus', 'f5': 'f5_tmsh', 'bigip': 'f5_tmsh', 'a10': 'a10',
    'linux': 'linux', 'ubuntu': 'linux', 'centos': 'linux', 'redhat': 'linux', 'debian': 'linux',
    'generic_termserver': 'generic_termserver', 'terminal_server': 'generic_termserver'
}

# 优化: 抽象厂商配置（新增 requires_enable 和 init_commands）
DEVICE_VENDOR_CONFIGS = {
    # Cisco厂商设备 (需要Enable)
    'cisco': {
        'timeout': 25, 'banner_timeout': 15, 'auth_timeout': 10, 'global_delay_factor': 1, 'conn_timeout': 10,
        'requires_enable': True,
    },
    
    # Huawei厂商设备 (不需要Enable，需要关闭分页)
    'huawei': {
        'timeout': 30, 'banner_timeout': 20, 'auth_timeout': 15, 'global_delay_factor': 2, 'conn_timeout': 15,
        'requires_enable': False,
        'init_commands': ['screen-length 0 temporary']
    },
    
    # Juniper厂商设备 (不需要Enable，需要关闭分页)
    'juniper': {
        'timeout': 35, 'banner_timeout': 25, 'auth_timeout': 15, 'global_delay_factor': 2, 'conn_timeout': 15,
        'requires_enable': False,
        'init_commands': ['set cli screen-length 0']
    },
    
    # HP/H3C厂商设备 (Comware不需要，Procurve需要)
    'hp': {
        'timeout': 25, 'banner_timeout': 15, 'auth_timeout': 10, 'global_delay_factor': 1, 'conn_timeout': 10,
        'requires_enable': True, # 默认为 True
    },
    
    # Fortinet厂商设备 (不需要Enable，需要设置终端)
    'fortinet': {
        'timeout': 30, 'banner_timeout': 20, 'auth_timeout': 15, 'global_delay_factor': 2, 'conn_timeout': 15,
        'requires_enable': False, 'use_keys': False, 'allow_agent': False,
        'init_commands': ['config system console', 'set output standard', 'end'] # Fortinet需要 send_config_set
    },
    
    # PaloAlto厂商设备 (不需要Enable，需要连接后延迟)
    'paloalto': {
        'timeout': 45, 'banner_timeout': 30, 'auth_timeout': 20, 'global_delay_factor': 3, 'conn_timeout': 20,
        'requires_enable': False, 'use_keys': False, 'allow_agent': False,
        'post_connect_sleep': 2 # 特殊处理：连接后等待
    },
    
    # Dell/Extreme/Ruckus (需要Enable)
    'dell': {'requires_enable': True, **{'timeout': 30, 'global_delay_factor': 1, 'conn_timeout': 10}},
    'extreme': {'requires_enable': True, **{'timeout': 30, 'global_delay_factor': 2, 'conn_timeout': 15}},
    'ruckus': {'requires_enable': True, **{'timeout': 30, 'global_delay_factor': 2, 'conn_timeout': 15}},
    
    # Mikrotik设备 (不需要Enable，需要连接后延迟)
    'mikrotik': {
        'timeout': 30, 'banner_timeout': 20, 'auth_timeout': 15, 'global_delay_factor': 3, 'conn_timeout': 15,
        'requires_enable': False,
        'post_connect_sleep': 1
    },
    
    # Linux设备 (不需要Enable，需要设置TERM)
    'linux': {
        'timeout': 30, 'global_delay_factor': 1, 'conn_timeout': 10, 'requires_enable': False,
        'init_commands': ['export TERM=vt100']
    },
    
    # 默认配置
    'default': {
        'timeout': 30, 'banner_timeout': 15, 'auth_timeout': 10, 'fast_cli': False, 
        'session_timeout': 60, 'global_delay_factor': 1, 'conn_timeout': 10,
        'requires_enable': False
    }
}

# ==============================================================================
# 辅助函数
# ==============================================================================

def thread_initializer() -> None:
    """线程初始化（解决编码问题）"""
    import encodings.idna
    encodings.idna.__name__

def sanitize_filename(name: str) -> str:
    """生成安全文件名"""
    return re.sub(r'[\\/*?:"<>|]', '', name).strip()[:60]

def normalize_device_type(device_type: str) -> str:
    """智能设备类型标准化"""
    original_type = device_type.lower().strip()
    if original_type in SUPPORTED_DEVICE_TYPES:
        return original_type
    if original_type in DEVICE_TYPE_ALIASES:
        mapped_type = DEVICE_TYPE_ALIASES[original_type]
        if mapped_type in SUPPORTED_DEVICE_TYPES:
            return mapped_type
    for supported_type in SUPPORTED_DEVICE_TYPES:
        if original_type in supported_type or supported_type in original_type:
            return supported_type
    return original_type

def get_device_vendor(device_type: str) -> str:
    """根据设备类型获取厂商"""
    device_type = device_type.lower()
    vendor_mapping = {
        'cisco': ['cisco_', 'ios', 'nxos', 'asa', 'wlc', 'xe', 'xr'],
        'huawei': ['huawei', 'vrp'],
        'juniper': ['juniper', 'junos', 'screenos'],
        'hp': ['hp_', 'aruba', 'comware', 'procurve', 'h3c'],
        'fortinet': ['fortinet'], 'paloalto': ['paloalto', 'panos'],
        'dell': ['dell_'], 'extreme': ['extreme'],
        'ruckus': ['ruckus', 'brocade', 'fastiron'],
        'mikrotik': ['mikrotik'], 'alcatel': ['alcatel', 'nokia'],
        'avaya': ['avaya'], 'f5': ['f5_'], 'a10': ['a10'],
        'linux': ['linux', 'ubuntu', 'centos']
    }
    for vendor, patterns in vendor_mapping.items():
        if any(pattern in device_type for pattern in patterns):
            return vendor
    return 'default'

def get_device_config(device_type: str) -> Dict[str, Any]:
    """获取设备特定配置"""
    vendor = get_device_vendor(device_type)
    return DEVICE_VENDOR_CONFIGS.get(vendor, DEVICE_VENDOR_CONFIGS['default']).copy()

def validate_device_data(device: Dict[str, str], row_idx: int) -> None:
    """验证设备数据完整性"""
    required = ['host', 'device_type']
    if missing := [f for f in required if not device.get(f)]:
        raise ValueError(f"Row {row_idx} 缺失字段: {', '.join(missing)}")
    
    normalized_type = normalize_device_type(device['device_type'])
    if normalized_type not in SUPPORTED_DEVICE_TYPES:
        # 优化: 使用 tqdm.write() 输出警告
        tqdm.write(f"[WARN] Row {row_idx} ({device['host']}): 未知设备类型 '{device['device_type']}' -> '{normalized_type}', 将使用默认配置", file=sys.stderr)

def load_excel(excel_file: str, sheet_name: str = 'Sheet1') -> List[Dict[str, str]]:
    """加载Excel设备清单"""
    devices = []
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        sheet = wb[sheet_name]
        
        headers = [str(cell.value).lower().strip() for cell in sheet[1]]
        required = ['host', 'device_type']
        if missing := [f for f in required if f not in headers]:
             raise ValueError(f"缺少必要列: {', '.join(missing)}")

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # 跳过空行
                continue
            device = {headers[i]: str(cell).strip() if cell else "" for i, cell in enumerate(row)}
            validate_device_data(device, row_idx)
            
            original_type = device['device_type']
            device['device_type'] = normalize_device_type(original_type)
            device['original_type'] = original_type 
            
            devices.append(device)
            
        return devices
    except Exception as e:
        print(f"Excel处理失败: {str(e)}")
        sys.exit(1)
    finally:
        if wb:
            wb.close()

def log_error(ip: str, error: str) -> None:
    """
    核心优化: 使用 tqdm.write() 代替 print() 确保进度条不被干扰
    安全记录错误日志，并使用 tqdm.write 避免进度条干扰
    """
    sanitized = re.sub(r'(password|secret)\s*=\s*\S+', r'\1=***', error, flags=re.I)
    log_line = f"{datetime.datetime.now():%Y-%m-%d %H:%M:%S} | {ip} | {sanitized}"
    
    with write_lock:
        try:
            with open("error.log", 'a', encoding='utf-8') as f:
                f.write(log_line + '\n')
            
            tqdm.write(f"{ip} [ERROR] {sanitized}", file=sys.stderr)
            
        except Exception as e:
            tqdm.write(f"FATAL ERROR writing log for {ip}: {str(e)}", file=sys.stderr)

# ==============================================================================
# 连接和执行
# ==============================================================================

def connect_device(device: Dict[str, str]) -> Optional[netmiko.BaseConnection]:
    """通用设备连接（支持所有netmiko设备）"""
    device_type = device['device_type']
    device_config = get_device_config(device_type)
    vendor = get_device_vendor(device_type)
    
    # 基础连接参数
    params = {
        'device_type': device_type, 'host': device['host'], 'username': device['username'],
        'password': device['password'], 'timeout': device_config['timeout'],
        'banner_timeout': device_config['banner_timeout'], 'auth_timeout': device_config['auth_timeout'],
        'global_delay_factor': device_config['global_delay_factor'], 'conn_timeout': device_config['conn_timeout'],
        'read_timeout_override': int(device.get('readtime', device_config['timeout'])),
        # 兼容性参数
        'fast_cli': device_config.get('fast_cli', False),
        'session_timeout': device_config.get('session_timeout', 60),
    }
    
    # 可选参数
    if device.get('secret'): params['secret'] = device['secret']
    if device.get('port'): params['port'] = int(device['port'])
    if 'use_keys' in device_config: params['use_keys'] = device_config['use_keys']
    if 'allow_agent' in device_config: params['allow_agent'] = device_config['allow_agent']
    if device_type.endswith('_telnet'):
        params.pop('use_keys', None); params.pop('allow_agent', None)
    
    # 调试日志配置
    if device.get('debug'):
        debug_dir = os.path.join("debug_logs", datetime.datetime.now().strftime('%Y%m%d'))
        os.makedirs(debug_dir, exist_ok=True)
        log_file = f"{sanitize_filename(device['host'])}_{uuid.uuid4().hex[:6]}.log"
        params['session_log'] = os.path.join(debug_dir, log_file)

    # 多重连接尝试 (使用 tqdm.write() 报告重试)
    max_retries = 2
    for attempt in range(max_retries + 1):
        try:
            conn = netmiko.ConnectHandler(**params)
            
            post_connection_setup(conn, device_type, vendor, device.get('secret'))
            
            return conn
            
        except (NetmikoTimeoutException, NetmikoAuthenticationException) as e:
            if attempt < max_retries:
                tqdm.write(f"[RETRY {attempt+1}] {device['host']}: {e.__class__.__name__}", file=sys.stderr)
                time.sleep(2 ** attempt)
                continue
            log_error(device['host'], f"{e.__class__.__name__}: {str(e)} (Type: {device.get('original_type', device_type)})")
        except Exception as e:
            if attempt < max_retries:
                tqdm.write(f"[RETRY {attempt+1}] {device['host']}: Connection error", file=sys.stderr)
                time.sleep(2 ** attempt)
                continue
            log_error(device['host'], f"连接异常: {str(e)} (Type: {device.get('original_type', device_type)})")
    
    return None

def post_connection_setup(conn: netmiko.BaseConnection, device_type: str, vendor: str, secret: Optional[str]) -> None:
    """
    简化连接后设置，抽象化配置
    处理 enable 模式、初始化命令和连接后延迟。
    """
    device_config = get_device_config(device_type)

    # 1. Enable 模式处理
    if secret and device_config.get('requires_enable', False):
        try:
            conn.enable()
        except (NetmikoAuthenticationException, NetmikoTimeoutException):
            pass # 预期失败，保持连接 (如设备已在特权模式)
        except Exception as e:
            tqdm.write(f"[WARN] {conn.host} Unexpected error during enable: {str(e)}. Continuing...", file=sys.stderr)
            pass
            
    # 2. 初始化命令
    if init_cmds := device_config.get('init_commands'):
        try:
            # Fortinet/H3C/Comware 的配置命令通常使用 send_config_set
            if vendor in ['fortinet', 'hp'] and conn.device_type in ['fortinet', 'hp_comware']:
                conn.send_config_set(init_cmds, cmd_verify=False)
            else:
                # 其他厂商发送单行命令
                for cmd in init_cmds:
                    # 对于 Huawei, Juniper, Linux 等，确保命令发送成功
                    conn.send_command(cmd, expect_string=r'[#>]', delay_factor=1)
        except Exception as e:
            tqdm.write(f"[WARN] {conn.host} Init commands failed: {str(e)}", file=sys.stderr)

    # 3. 连接后延迟
    if sleep_time := device_config.get('post_connect_sleep'):
        time.sleep(sleep_time)


def execute_commands(device: Dict[str, str], config_set: bool) -> Optional[str]:
    """通用命令执行（适配所有设备类型）"""
    device_type = device['device_type']
    
    try:
        cmds = [c.strip() for c in device.get('mult_command', '').split(';') if c.strip()]
        if not cmds:
            # 使用 tqdm.write() 输出警告
            tqdm.write(f"{device['host']} [WARN] 无有效命令", file=sys.stderr)
            return None

        if not (conn := connect_device(device)):
            return None

        with conn:
            vendor = get_device_vendor(device_type)
            
            # 获取设备主机名
            try:
                # 优先尝试使用 netmiko 的 base_prompt
                device['hostname'] = getattr(conn, 'base_prompt', 'unknown').split('(')[0].strip()
                if device['hostname'] == 'unknown' or not device['hostname']:
                     device['hostname'] = extract_hostname(conn, device_type, vendor)
            except:
                device['hostname'] = 'unknown'

            # 执行命令
            all_output = []
            if config_set:
                all_output.extend(execute_config_commands(conn, cmds, device_type, vendor))
            else:
                all_output.extend(execute_show_commands(conn, cmds, device_type, vendor))

            return "\n\n".join(all_output)
            
    except Exception as e:
        log_error(device['host'], f"执行异常 ({device.get('original_type', device_type)}): {str(e)}")
        return None

def extract_hostname(conn: netmiko.BaseConnection, device_type: str, vendor: str) -> str:
    """提取设备主机名（多厂商适配的回退逻辑）"""
    try:
        prompt = conn.find_prompt().strip()
        
        hostname_patterns = {
            'paloalto': [r'(\S+?)[@#>]', r'$(\S+?)$', r'(\S+)[@#>]'],
            'fortinet': [r'$(\S+?)$', r'(\S+?)[#>]', r'(\S+)-'],
            'juniper': [r'(\S+?)[@#>]', r'(\S+?)%'],
            'mikrotik': [r'$$(\S+?)$$', r'(\S+?)>'],
            'linux': [r'(\S+?)[@#$]', r'(\S+?):'],
            'f5': [r'$(\S+?)$', r'(\S+?)#'],
            'default': [r'(\S*?)([\w.-]+)[#<>@$]', r'(\S+?)[#>]', r'(\w+)']
        }
        
        patterns = hostname_patterns.get(vendor, hostname_patterns['default'])
        
        for pattern in patterns:
            match = re.search(pattern, prompt)
            if match:
                hostname = match.group(1)
                if hostname and len(hostname) > 1:
                    return hostname
                    
        return 'unknown'
    except:
        return 'unknown'

def execute_config_commands(conn: netmiko.BaseConnection, cmds: List[str], device_type: str, vendor: str) -> List[str]:
    """配置模式命令执行"""
    outputs = []
    
    try:
        # PaloAlto/Fortinet 的配置命令通常在 enable 模式下单独发送
        if vendor in ['paloalto', 'fortinet']:
            for cmd in cmds:
                # Fortinet 配置命令必须使用 send_config_set
                if vendor == 'fortinet':
                    output = conn.send_config_set([cmd], cmd_verify=False)
                else:
                    output = conn.send_command(cmd, expect_string=r'[#>$]', delay_factor=2)
                outputs.append(f"Config Command: {cmd}\n{output}")
        else:
            # 标准配置模式
            output = conn.send_config_set(cmds, cmd_verify=False)
            outputs.append(output)
            
    except Exception as e:
        outputs.append(f"Config execution error: {str(e)}")
    
    return outputs

def execute_show_commands(conn: netmiko.BaseConnection, cmds: List[str], device_type: str, vendor: str) -> List[str]:
    """查看命令执行"""
    outputs = []
    
    # 厂商特定的延迟因子
    delay_factors = {'huawei': 2, 'paloalto': 3, 'fortinet': 2, 'juniper': 2, 'ruckus': 2, 'extreme': 2, 'mikrotik': 3, 'default': 1}
    delay_factor = delay_factors.get(vendor, 1)
    
    for cmd in cmds:
        try:
            output = conn.send_command(cmd, cmd_verify=False, delay_factor=delay_factor)
            outputs.append(f"Command: {cmd}\n{output}")
            
        except Exception as e:
            # 细化命令执行失败时的日志
            log_error(conn.host, f"Command execution failed: {cmd} - {str(e)}")
            outputs.append(f"Command: {cmd}\nError: {str(e)}")
    
    return outputs

def save_result(ip: str, hostname: str, output: str, dest_path: str, device_type: str = '', original_type: str = '') -> None:
    """保存执行结果"""
    date_str = datetime.datetime.now().strftime('%Y%m%d')
    output_dir = os.path.join(dest_path, f"result_{date_str}")
    os.makedirs(output_dir, exist_ok=True)

    vendor = get_device_vendor(device_type)
    filename = f"{ip}_{sanitize_filename(hostname)}.txt"
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    content = f"""=== 设备信息 ===

IP地址: {ip}
主机名: {sanitize_filename(hostname)}
设备类型: {device_type}
原始类型: {original_type}
厂商: {vendor}
执行时间: {timestamp}

=== 执行结果 ===
{output}"""

    with write_lock: # 在写入文件时使用锁
        try:
            with tempfile.NamedTemporaryFile(
                mode='w',
                encoding='utf-8',
                delete=False,
                dir=output_dir
            ) as tmp_file:
                tmp_file.write(content)
                tmp_path = tmp_file.name
            
            os.rename(tmp_path, os.path.join(output_dir, filename))
        except OSError as e:
            log_error(ip, f"文件保存失败: {str(e)}")

def batch_execute(
    devices: List[Dict[str, str]],
    config_set: bool,
    max_workers: int = DEFAULT_THREADS,
    destination: str = './'
) -> None:
    """批量执行"""
    success_count = 0
    try:
        with ThreadPoolExecutor(
            max_workers=max_workers,
            initializer=thread_initializer
        ) as executor:
            futures = {executor.submit(execute_commands, dev, config_set): dev for dev in devices}
            # 确保 tqdm 进度条只在终端运行时显示
            progress = tqdm(total=len(devices), desc="执行进度", unit="台", disable=(not sys.stdout.isatty()))

            try:
                for future in as_completed(futures):
                    dev = futures[future]
                    try:
                        if (result := future.result()) is not None:
                            save_result(
                                dev['host'], dev.get('hostname', 'unknown'), result, 
                                destination, dev['device_type'], dev.get('original_type', dev['device_type'])
                            )
                            success_count += 1
                    except Exception as e:
                        # 错误已经在 log_error 中处理
                        pass 
                    finally:
                        progress.update(1)
                progress.close()
                tqdm.write(f"\n**完成**: 成功 {success_count}/{len(devices)} 台设备")
            except KeyboardInterrupt:
                progress.close()
                executor.shutdown(wait=False, cancel_futures=True)
                raise
    except KeyboardInterrupt:
        sys.exit(0)

def list_supported_devices() -> None:
    """显示所有支持的设备类型"""
    print(f"**支持的设备类型总数**: {len(SUPPORTED_DEVICE_TYPES)}\n")
    
    vendor_devices = {}
    for device_type in sorted(SUPPORTED_DEVICE_TYPES):
        vendor = get_device_vendor(device_type)
        if vendor not in vendor_devices:
            vendor_devices[vendor] = []
        vendor_devices[vendor].append(device_type)
    
    for vendor, devices in sorted(vendor_devices.items()):
        print(f"**{vendor.upper()}** ({len(devices)} 种):")
        for device in sorted(devices):
            print(f"  - {device}")
        print()

def parse_args() -> argparse.Namespace:
    """命令行参数解析"""
    parser = argparse.ArgumentParser(
        description="**网络设备批量管理工具 v5.0** - 支持所有netmiko设备",
        add_help=False,
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument('-i', '--input', help='设备清单Excel路径')
    parser.add_argument('-t', '--threads', type=int, default=DEFAULT_THREADS, help=f'并发线程数 (默认: {DEFAULT_THREADS})')
    parser.add_argument('-cs', '--config_set', action='store_true', help='使用配置模式发送命令')
    parser.add_argument('-d', '--destination', default='./', help='结果保存路径')
    parser.add_argument('--debug', action='store_true', help='启用调试日志')
    parser.add_argument('-s', '--sheet', default='Sheet1', help='Excel工作表名称')
    parser.add_argument('--list-devices', action='store_true', help='列出所有支持的设备类型')

    if '--version' in sys.argv or '-V' in sys.argv:
        print("Python", platform.python_version(), platform.platform())
        print("Netmiko", netmiko.__version__)
        sys.exit(0)

    # 处理帮助信息 (保留用户要求的结构)
    if '--help' in sys.argv or '-h' in sys.argv:
        
        help_text = f"""
**网络设备批量管理工具 v5.0 - 全设备支持版本**

**特性**:
- **支持 {len(SUPPORTED_DEVICE_TYPES)} 种设备类型** (所有netmiko支持的设备)
- **智能设备类型识别** (支持别名和模糊匹配)
- **厂商特定优化配置** (针对不同厂商调优)
- **自动重试机制** (连接失败自动重试)
- **并发执行** (多线程提高效率)

**使用方法**:
  python net_cli_manager.py -i <设备清单.xlsx> [选项]

**参数说明**:
  -i, --input        必需  Excel文件路径  
  -t, --threads      可选  并发线程数 (默认: {DEFAULT_THREADS})
  -cs, --config_set  可选  使用配置模式
  -d, --destination  可选  结果保存路径
  -s, --sheet        可选  Excel工作表名
  --debug            可选  启用详细日志
  --list-devices     可选  显示支持的设备类型

**Excel格式**:
| host        | username | password | device_type  | secret | port | mult_command               |
|-------------|----------|----------|--------------|--------|------|----------------------------|
| 192.168.1.1 | admin    | pass123  | cisco_ios    | enable | 22   | show version;sh clock      |
| 192.168.1.2 | admin    | pass123  | huawei       |        | 22   | disp version;disp clock    |
| 192.168.1.3 | admin    | pass123  | paloalto     |        | 22   | show system info           |

**示例**:
  # 基本使用
  python net_cli_manager.py -i devices.xlsx
  
  # 配置模式 + 调试
  python net_cli_manager.py -i devices.xlsx -cs --debug
  
  # 查看支持的设备类型
  python net_cli_manager.py --list-devices
"""
        print(help_text)
        sys.exit(0)

    return parser.parse_args()

def main() -> None:
    """主入口"""
    args = parse_args()
    
    if args.list_devices:
        list_supported_devices()
        return
    
    if not args.input:
        print("**错误**: 必须指定输入文件 (-i)")
        sys.exit(1)
        
    if not os.path.exists(args.input):
        print(f"**错误**: 文件不存在 [{args.input}]")
        sys.exit(1)
        
    try:
        devices = load_excel(args.input, args.sheet)
        
        if args.debug:
            for device in devices:
                device['debug'] = True
        
        # 显示设备统计信息
        device_stats = {}; vendor_stats = {}
        for device in devices:
            device_type = device['device_type']
            original_type = device.get('original_type', device_type)
            vendor = get_device_vendor(device_type)
            key = f"{original_type} -> {device_type}" if original_type != device_type else device_type
            device_stats[key] = device_stats.get(key, 0) + 1
            vendor_stats[vendor] = vendor_stats.get(vendor, 0) + 1
        
        print(f"**成功加载设备**: {len(devices)} 台 (工作表: {args.sheet})")
        print(f"**网络设备管理工具** - 支持 {len(SUPPORTED_DEVICE_TYPES)} 种设备类型 (最大线程: {args.threads})")
        
        print("\n**厂商分布**:")
        for vendor, count in sorted(vendor_stats.items()):
            print(f"  - **{vendor.upper()}**: {count} 台")
        
        print("\n**设备类型详情**:")
        for device_type, count in sorted(device_stats.items()):
            print(f"  - {device_type}: {count} 台")
        
        print()
        batch_execute(devices, args.config_set, args.threads, args.destination)
        
    except KeyboardInterrupt:
        print("\n**用户终止**")
        sys.exit(0)

if __name__ == "__main__":
    main()
    
